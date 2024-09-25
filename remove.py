"""

To remove datas from excel file use this script change the file name & edit with your own data in txt_file

"""





import openpyxl
 
file_path = 'Non_Kernel_Patching_data.xlsx'
workbook = openpyxl.load_workbook(file_path)
 
txt_file = 'data_remove.txt'
with open(txt_file, 'r') as file:
    data_remove = [line.strip() for line in file.readlines()]
 
removed_data_log = []
 
data_removed = False
 
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in data_remove:
                removed_data_log.append(f"Sheet: {sheet_name}, Cell: {cell.coordinate}, Value: {cell.value}")
                cell.value = ''  
                data_removed = True
 
workbook.save(file_path)
 
if data_removed:
    print("Matching data removed from all sheets and saved successfully.")
    print("Log of removed data:")
    for entry in removed_data_log:
        print(entry)
else:
    print("No matching data found in the workbook.")
